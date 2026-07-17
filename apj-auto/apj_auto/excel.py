# -*- coding: utf-8 -*-
"""APJ発注書 xlsx の生成。

レイアウトは apj-order-app の ExcelJS 実装を openpyxl に移植したもの。
シート名: APJ発注書 / ファイル名: APJ発注書_YYYYMMDD.xlsx
"""

import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .transform import OUTPUT_HEADERS

COL_WIDTHS = [45, 30, 6, 16, 12, 10, 18, 52, 16, 16, 18]
N_COLS = len(OUTPUT_HEADERS)  # 11

FONT_NAME = "Meiryo"

_MEDIUM_BLACK = Side(style="medium", color="FF000000")
_THIN_GRAY = Side(style="thin", color="FFCCCCCC")


def _font(bold=False, size=10, color=None):
    return Font(name=FONT_NAME, size=size, bold=bold,
                color=color if color else None)


def build_order_xlsx(rows: list, out_dir: Path, today: datetime.date) -> Path:
    """変換済み行データから発注書xlsxを生成し、ファイルパスを返す。"""
    date_jp = f"{today.year}年{today.month:02d}月{today.day:02d}日"

    wb = Workbook()
    ws = wb.active
    ws.title = "APJ発注書"
    ws.sheet_view.showGridLines = False

    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # Row 1: 見出し + 発注日（下線のみ）
    ws.cell(r, 1, "アートプリントジャパン 御中")
    ws.cell(r, 1).font = _font(bold=True, size=14)
    ws.cell(r, 1).border = Border(bottom=_MEDIUM_BLACK)
    ws.cell(r, 2, f"発注日：{date_jp}")
    ws.cell(r, 2).font = _font()
    ws.cell(r, 2).alignment = Alignment(horizontal="right", vertical="bottom")
    ws.cell(r, 2).border = Border(bottom=_MEDIUM_BLACK)
    ws.row_dimensions[r].height = 28

    r += 1  # Row 2: 空行
    ws.row_dimensions[r].height = 8

    r += 1  # Row 3: 発注者
    ws.cell(r, 1, "発注者：有限会社 EL COMMUN　担当：真田").font = _font(bold=True)
    ws.row_dimensions[r].height = 17

    r += 1  # Row 4: 住所
    ws.cell(r, 1, "〒468-0007 愛知県名古屋市天白区植田本町2丁目1006番地").font = \
        _font(color="FF444444")
    ws.row_dimensions[r].height = 17

    r += 1  # Row 5: TEL/FAX
    ws.cell(r, 1, "TEL：052-807-0299　FAX：052-685-5116").font = _font(color="FF444444")
    ws.row_dimensions[r].height = 17

    r += 1  # 空行（区切り）
    ws.row_dimensions[r].height = 6

    r += 1  # 挨拶文
    ws.cell(r, 1, "お手数をおかけ致しますが、下記の直送手配を宜しくお願い申し上げます。").font = _font()
    ws.row_dimensions[r].height = 17

    r += 1  # 送り状連絡依頼
    ws.cell(r, 1, "発送後は送り状番号をメールにてご連絡いただけますと幸いです。").font = _font()
    ws.row_dimensions[r].height = 17

    r += 1  # お願い事項
    ws.cell(r, 1, "＊配送料は記載しないでください　＊お届け日指定が空欄の注文は 日時の指定は不要です").font = \
        _font(bold=True)
    ws.row_dimensions[r].height = 17

    r += 1  # 空行
    ws.row_dimensions[r].height = 8

    r += 1  # 列ヘッダー（薄グレー背景、下線のみ）
    header_fill = PatternFill(fill_type="solid", fgColor="FFF3F4F6")
    for c, h in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(r, c, h)
        cell.fill = header_fill
        cell.font = _font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=_MEDIUM_BLACK)
    ws.row_dimensions[r].height = 20

    # データ行（下線のみ、最終行は太線）
    for idx, row in enumerate(rows):
        r += 1
        is_last = idx == len(rows) - 1
        border = Border(bottom=_MEDIUM_BLACK if is_last else _THIN_GRAY)
        for c, h in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(r, c, row.get(h, ""))
            cell.font = _font()
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if h in ("郵便番号", "電話番号"):
                cell.number_format = "@"  # 先頭0保持のため文字列扱い
        ws.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(r, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 18

    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"APJ発注書_{today.strftime('%Y%m%d')}.xlsx"
    path = out_dir / fname
    wb.save(path)
    return path
