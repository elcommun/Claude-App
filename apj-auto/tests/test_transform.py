# -*- coding: utf-8 -*-
"""変換ロジックのテスト（apj-order-app の挙動と一致することの確認）。

実行: cd apj-auto && python -m pytest tests/ -v
      （pytest が無ければ python tests/test_transform.py でも可）
"""

import datetime
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from apj_auto.transform import (  # noqa: E402
    convert_row,
    parse_goq_csv,
    process_options,
    process_product_name,
)


def test_product_name_apj_prefix():
    assert process_product_name("【APJ】FIT FRAME【610mm×915mm】") == \
        "FIT FRAME【610mm×915mm】"


def test_product_name_trailing_text_removed():
    # 最初の【】以降の付帯文言は除去され、サイズ表記だけが残る
    assert process_product_name(
        "【APJ】FIT FRAME【610mm×915mm】ポスターフレーム 額縁"
    ) == "FIT FRAME【610mm×915mm】"


def test_product_name_apj_code_and_coupon():
    assert process_product_name(
        "【APJ】＼P2倍＋クーポン／FIT FRAME(apj-1000123)【B2】"
    ) == "FIT FRAME【B2】"


def test_product_name_real_sample_20260717():
    # 2026-07-17 の実受注CSV（GoQダウンロード）と apj-order-app 生成xlsxの
    # 実データペアで検証済みのパターン（セル値完全一致を確認済み）
    s = ("【APJ】FIT FRAME【610mm×915mm】（61cm×91.5cm）ポスターサイズ（額縁内寸）"
         "アルミ製額縁／アルミポスターフレーム全4色30サイズ・オーダーメイド も可能！"
         "【受注生産】【発送までに 約7-10営業日】【「APJ」以外の商品と同送不可】"
         "(apj-af61_91)")
    assert process_product_name(s) == "FIT FRAME【610mm×915mm】"
    s2 = ("【APJ】FIT FRAME【348×424mm：四つ切サイズ】（額縁内寸）アルミ製額縁／"
          "アルミポスターフレーム全4色30サイズ・オーダーメイド も可能！【受注生産】"
          "【発送までに 約7-10営業日】【「APJ」以外の商品と同送不可】(apj-af34_42)")
    assert process_product_name(s2) == "FIT FRAME【348×424mm：四つ切サイズ】"


def test_options_real_sample_20260717():
    s = ("カラー:ブラック-0020195326\n"
         "【長期商品欠品の場合】:1-2営業日中にご連絡します\n"
         "【日曜 / 祝日着 指定不可】:○了解の上ご購入\n"
         "発送までに約7-10営業日")
    assert process_options(s) == "ブラック-0020195326"


def test_product_name_no_size():
    assert process_product_name("【APJ】ポスターフレーム A4") == "ポスターフレーム A4"


def test_product_name_empty():
    assert process_product_name("") == ""


def test_options_color_prefix():
    assert process_options("カラー:ホワイト") == "ホワイト"


def test_options_multiline_and_note():
    s = "カラー:ブラック【長期商品のためお届けまで2週間】\n【長期商品につき…】"
    assert process_options(s) == "ブラック"


def test_options_hassou_made():
    assert process_options("フレームカラー=ナチュラル発送までに1週間") == "ナチュラル"


def test_options_design_prefix():
    assert process_options("デザイン:シルバー") == "シルバー"


def test_options_darkblue():
    assert process_options("カラー:ダークブルー") == "ブルー"


def test_convert_row_drops_biko():
    row = ["【APJ】FIT FRAME【A3】", "カラー:白", "2", "山田太郎",
           "468-0007", "愛知県", "名古屋市天白区", "植田本町2-1006",
           "052-807-0299", "", "", "備考メモ"]
    out = convert_row(row)
    assert out["商品名"] == "FIT FRAME【A3】"
    assert out["個数"] == "2"
    assert "備考" not in out
    assert out["郵便番号"] == "468-0007"


def test_parse_goq_csv_cp932():
    csv_text = (
        "商品名(削除),項目・選択肢(削除),個数,氏名,郵便番号,都道府県,都市区,町以降,"
        "電話番号(削除),お届け日指定(削除),お届け時間帯(削除),備考(削除)\r\n"
        "【APJ】FIT FRAME【B2】,カラー:黒,1,テスト花子,001-0001,北海道,札幌市,中央区1-1,"
        "011-111-1111,,,\r\n"
    )
    rows = parse_goq_csv(csv_text.encode("cp932"))
    assert len(rows) == 1
    assert rows[0]["商品名"] == "FIT FRAME【B2】"
    assert rows[0]["項目・選択肢"] == "黒"
    assert rows[0]["郵便番号"] == "001-0001"  # 先頭0保持


def test_excel_generation():
    from apj_auto.excel import build_order_xlsx
    from openpyxl import load_workbook

    rows = [convert_row(["【APJ】FIT FRAME【A3】", "カラー:白", "2", "山田太郎",
                         "060-0001", "北海道", "札幌市", "中央区1-1",
                         "090-1111-2222", "7/20", "午前中", "メモ"])]
    with tempfile.TemporaryDirectory() as d:
        path = build_order_xlsx(rows, Path(d), datetime.date(2026, 7, 16))
        assert path.name == "APJ発注書_20260716.xlsx"
        wb = load_workbook(path)
        ws = wb["APJ発注書"]
        assert ws.cell(1, 1).value == "アートプリントジャパン 御中"
        assert ws.cell(1, 2).value == "発注日：2026年07月16日"
        assert ws.cell(11, 1).value == "商品名"       # ヘッダー行
        assert ws.cell(12, 1).value == "FIT FRAME【A3】"
        assert ws.cell(12, 5).value == "060-0001"     # 郵便番号は文字列
        assert ws.cell(12, 5).number_format == "@"
        # 備考列が無い(11列)こと
        assert ws.cell(11, 11).value == "お届け時間帯"
        assert ws.cell(11, 12).value is None


def test_business_day():
    from apj_auto.guard import is_business_day
    assert is_business_day(datetime.date(2026, 7, 16)) is True    # 木曜
    assert is_business_day(datetime.date(2026, 7, 18)) is False   # 土曜
    assert is_business_day(datetime.date(2026, 7, 20)) is False   # 海の日(祝)
    assert is_business_day(datetime.date(2026, 1, 1)) is False    # 元日


def test_mail_body_matches_template():
    # 依頼された定型文と完全一致すること（全角の／・～を含む）
    from apj_auto.mailer import MAIL_BODY
    assert "神代様 斎藤様" in MAIL_BODY
    assert "　（営業時間：平日／AM9時～PM18時）" in MAIL_BODY
    assert "　E-mail：ec@elcommun.co.jp" in MAIL_BODY


def test_run_state_review_step():
    from apj_auto.guard import RunState
    with tempfile.TemporaryDirectory() as d:
        day = datetime.date(2026, 7, 17)
        s = RunState(Path(d), day)
        s.mark("fetched", order_ids=["B1"])
        s.mark("xlsx", xlsx_path="/tmp/x.xlsx")
        s.mark("review_sent")
        s2 = RunState(Path(d), day)
        assert s2.step_done("review_sent")
        assert not s2.step_done("mail_sent")
        assert not s2.completed  # 承認(--approve)まで完了扱いにしない


def test_run_state_resume():
    from apj_auto.guard import RunState
    with tempfile.TemporaryDirectory() as d:
        day = datetime.date(2026, 7, 16)
        s = RunState(Path(d), day)
        assert not s.completed
        s.mark("fetched", order_ids=["A1", "A2"])
        s.mark("mail_sent")
        # 再読込 → メール送信済みが引き継がれる（再送防止）
        s2 = RunState(Path(d), day)
        assert s2.step_done("mail_sent")
        assert s2.data["order_ids"] == ["A1", "A2"]
        assert not s2.completed
        s2.finish("done")
        assert RunState(Path(d), day).completed


if __name__ == "__main__":
    fails = 0
    for name, fn in sorted(globals().items()):
        if name.startswith("test_") and callable(fn):
            try:
                fn()
                print(f"OK   {name}")
            except AssertionError as e:
                fails += 1
                print(f"FAIL {name}: {e}")
    sys.exit(1 if fails else 0)
