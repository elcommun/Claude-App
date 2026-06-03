#!/usr/bin/env python3
"""
update_stockout.py — 欠品ExcelをパースしてSTOCKOUTをindex.htmlに書き込む

使い方:
  python item-data/update_stockout.py <Excelファイル>

Excel形式: A列=品番、B列=欠品日（1行目がヘッダー「品番」「欠品日」）
"""
import sys
import os
import re
import json
from openpyxl import load_workbook

# ── スクリプトからindex.htmlへのパス ──────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INDEX_HTML = os.path.join(SCRIPT_DIR, '..', 'ranking', 'index.html')


def load_stockout(html_content):
    """index.html から let STOCKOUT = {...}; を読み込んで dict を返す"""
    m = re.search(r'let STOCKOUT = (\{[^;]+\});', html_content)
    if not m:
        raise ValueError("index.html に let STOCKOUT = が見つかりません")
    return json.loads(m.group(1))


def process_excel(filepath):
    """欠品ExcelをパースしてSTOCKOUTデータ dict を返す"""
    print(f"  読み込み: {filepath}")
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Excelファイルを開けません: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    stockout = {}
    count = 0
    for row in rows:
        if not row or row[0] is None or row[1] is None:
            continue
        code = str(row[0]).strip()
        date = str(row[1]).strip()
        # ヘッダー行をスキップ
        if not code or code == '品番' or not date or date == '欠品日':
            continue
        stockout[code] = date
        count += 1

    print(f"  読み込み完了: {count} 件")
    return stockout


def write_stockout(html_content, stockout):
    """
    index.html の以下の2箇所を更新:
      1. let STOCKOUT = {...};
      2. clearStockout() 内の STOCKOUT={...};renderStockoutList()
    """
    json_str = json.dumps(stockout, ensure_ascii=False, separators=(',', ':'))

    # 1. let STOCKOUT = {...};
    new_content, count1 = re.subn(
        r'let STOCKOUT = \{[^;]+\};',
        f'let STOCKOUT = {json_str};',
        html_content,
        count=1,
    )
    if count1 == 0:
        raise ValueError("index.html の let STOCKOUT = が置換できませんでした")

    # 2. clearStockout() 内の STOCKOUT={...};if(selectedCategory)
    new_content, count2 = re.subn(
        r'(STOCKOUT=)\{[^}]*(?:\}[^}]*)*?\}(;if\(selectedCategory\))',
        rf'\g<1>{json_str}\2',
        new_content,
        count=1,
    )
    if count2 == 0:
        raise ValueError("index.html の clearStockout() 内の STOCKOUT={...} が置換できませんでした")

    return new_content


def main():
    if len(sys.argv) != 2:
        print("使い方: python item-data/update_stockout.py <Excelファイル>")
        sys.exit(1)

    excel_file = sys.argv[1]
    if not os.path.exists(excel_file):
        print(f"エラー: ファイルが見つかりません: {excel_file}")
        sys.exit(1)

    print(f"index.html を読み込み中: {INDEX_HTML}")
    with open(INDEX_HTML, 'r', encoding='utf-8') as f:
        html_content = f.read()

    existing_stockout = load_stockout(html_content)
    print(f"  既存データ: {len(existing_stockout)} 件")

    new_data = process_excel(excel_file)
    if not new_data:
        print("警告: 有効な欠品データがありませんでした")
        sys.exit(1)

    # マージ（新データ優先）
    merged = dict(existing_stockout)
    merged.update(new_data)
    print(f"  マージ後: {len(merged)} 件")

    print("index.html を更新中...")
    new_html = write_stockout(html_content, merged)

    with open(INDEX_HTML, 'w', encoding='utf-8') as f:
        f.write(new_html)

    print(f"完了: {len(new_data)} 件の欠品データを更新しました（合計 {len(merged)} 件）")


if __name__ == '__main__':
    main()
