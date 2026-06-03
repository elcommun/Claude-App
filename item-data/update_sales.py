#!/usr/bin/env python3
"""
update_sales.py — 販売明細ExcelをパースしてPRELOADEDをindex.htmlに書き込む

使い方:
  python item-data/update_sales.py <Excelファイル1> [<Excelファイル2> ...]
"""
import sys
import os
import re
import json
import datetime
from openpyxl import load_workbook

# ── スクリプトからindex.htmlへのパス ──────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
INDEX_HTML = os.path.join(SCRIPT_DIR, '..', 'ranking', 'index.html')

# ── カテゴリマッピング（index.html と同等） ────────────────────────────────────
CATEGORY_ORDER = [
    "ブックカバー","レターセット","保冷バッグ","ペンケース","ブックマーク","メッセージカード",
    "ぽち袋・金封","ポーチ・ケース","ノート","ステッカー・シール","メガネケース","マスキングテープ",
    "スタンプ","メモ","マウスパッド","ふせん","ポスター","グリーティングカード","下敷き","ボールペン",
    "トートバッグ","時計","ランチボックス","キーリング","エコバッグ","箸セット","タグ","クリアファイル",
    "食器","他カバー系","ハンカチクロス","コースター他","ブローチ","アウトドア","ブランケット",
    "ライトスポーツ","その他",
]

CAT_MAP = {
    'ランチ関係': 'ランチボックス',
    '付箋': 'ふせん',
    'シール・ステッカー': 'ステッカー・シール',
    'カード': 'グリーティングカード',
    '輸入カード': 'グリーティングカード',
    '他': 'ブックカバー',
    '別注': 'ブックカバー',
    '他紙製品': 'マスキングテープ',
    'デスクトップ': 'マウスパッド',
    'ステーショナリー': 'ペンケース',
    'ポーチ': 'ポーチ・ケース',
    'トート': 'トートバッグ',
    'ペン': 'ボールペン',
    '金封': 'ぽち袋・金封',
    'ラッピング': 'ぽち袋・金封',
    'バッグ類': '他カバー系',
    'ショルダー': '他カバー系',
    'キャリング': '他カバー系',
    'リュック': '他カバー系',
    'ファッション小物': '他カバー系',
    'プレート・ボウル・茶碗': '食器',
    'マグ・カップ・グラス・タンブラー': '食器',
    'がまくんとかえるくん': 'グリーティングカード',
    'エリック・カール': 'グリーティングカード',
    'レオ・レオニ': 'グリーティングカード',
    'ハンカチクロス': 'ハンカチクロス',
    'コースター他': 'コースター他',
    'ライトスポーツ': 'その他',
    'ピクニックシート': 'その他',
    'ブランケット': 'その他',
    'ランプ・ライト': 'その他',
    'インテリア': 'その他',
    'インテリア　生活雑貨': 'その他',
    'ツール': 'その他',
    'ボール': 'その他',
    'スマイル': 'その他',
    'おもちゃ': 'その他',
    '小物': 'その他',
    'アウトドアグッズ': 'その他',
    '生活雑貨＆他': 'その他',
    '他紙製品＆他': 'その他',
    '未分類': 'その他',
    '別注（仕入れ）': 'その他',
}

# コードのリネーム
CODE_REMAP = {'WCL-017WCL-018': 'WCL-018', 'B6-1J7F-DA1E': 'CPC-004'}
# 特定コードのカテゴリ強制割り当て
CODE_REMAP_CAT = {'WCL-018': 'ペンケース', 'CPC-004': '時計', 'WPK-101': 'その他'}
# 除外するコードプレフィックス
CODE_SKIP_PREFIXES = ['SLW-', 'DR-MM-']
# コードプレフィックスによるカテゴリ上書き
CODE_PREFIX_TO_CAT = [
    ('CRB-', 'アウトドア'), ('PCS-', 'アウトドア'), ('CLB-6', 'アウトドア'),
    ('AMP-ST', 'アウトドア'), ('STR-', 'アウトドア'), ('MST-', 'アウトドア'),
    ('LD-', 'アウトドア'), ('HOR-', 'アウトドア'),
    ('BKT-', 'ブランケット'),
    ('KIT-', 'ライトスポーツ'), ('SFD-', 'ライトスポーツ'),
    ('TWH-', 'ハンカチクロス'),
    ('GB-', 'ステッカー・シール'),
    ('MPD-', 'ステーショナリー'),
    ('TCL-6', 'ブローチ'), ('TCL-7', 'ブローチ'), ('MM-09', 'ブローチ'),
]


def get_code_cat(code):
    """CODE_REMAP_CAT → CODE_PREFIX_TO_CAT の順に優先カテゴリを返す。なければ None"""
    if code in CODE_REMAP_CAT:
        return CODE_REMAP_CAT[code]
    for prefix, cat in CODE_PREFIX_TO_CAT:
        if code.startswith(prefix):
            return cat
    return None


def map_category(cat):
    """raw カテゴリ名 → 正規カテゴリ名（index.html の mapCategory 相当）"""
    # 全角スペースを半角に正規化
    c = cat.replace('　', ' ')
    # 手帳・カレンダー系はそのまま通す
    if 'ダイアリー' in c or 'カレンダー' in c:
        return c
    if c in CAT_MAP:
        return CAT_MAP[c]
    if c in CATEGORY_ORDER:
        return c
    return 'その他'


def parse_date(val):
    """セル値を YYYY-MM-DD 文字列に変換。失敗したら空文字"""
    if val is None:
        return ''
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
        return s
    m = re.match(r'^(\d{4})/(\d{1,2})/(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f"{m.group(3)}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"
    m = re.match(r'^(\d{4})年(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    m = re.match(r'^令和(\d+)年(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f"{2018 + int(m.group(1))}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    m = re.match(r'^平成(\d+)年(\d{1,2})月(\d{1,2})日', s)
    if m:
        return f"{1988 + int(m.group(1))}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"
    return ''


def load_preloaded(html_path):
    """index.html から PRELOADED の JSON を読み込む"""
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()
    m = re.search(r'const PRELOADED=(\{.*?\});?\s*\n', content, re.DOTALL)
    if not m:
        raise ValueError("index.html に const PRELOADED= が見つかりません")
    return json.loads(m.group(1)), content


def process_excel(filepath):
    """
    Excelファイルを処理して monthly_data を返す。
    戻り値:
      monthly_data: { 'YYYY-MM': { 'by_store': {...}, 'products': {...}, 'date_range': [min,max] } }
      processed_yms: set of 'YYYY-MM' strings found
    """
    print(f"  読み込み: {filepath}")
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Excelファイルを開けません: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        print(f"  警告: データ行がありません ({filepath})")
        return {}, set()

    hdr = [str(h).strip() if h is not None else '' for h in rows[0]]

    def col(names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return -1

    date_col   = col(['販売日', '出荷日'])
    name_col   = col(['商品名'])
    code_col   = col(['商品番号'])
    price_col  = col(['税抜単価', '商品単価'])
    qty_col    = col(['数量', '個数'])
    store_col  = col(['店舗', '販売店舗'])
    scat_col   = col(['小分類'])
    sc_col     = col(['検索コード'])

    if date_col < 0:
        raise ValueError(f"列「販売日」または「出荷日」が見つかりません ({filepath})")
    if code_col < 0:
        raise ValueError(f"列「商品番号」が見つかりません ({filepath})")
    if qty_col < 0:
        raise ValueError(f"列「数量」または「個数」が見つかりません ({filepath})")

    monthly_data = {}
    processed_yms = set()

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue
        date_str = parse_date(row[date_col] if date_col < len(row) else None)
        if not date_str:
            continue
        ym = date_str[:7]

        raw_code = str(row[code_col]).strip() if code_col < len(row) and row[code_col] is not None else ''
        if not raw_code:
            continue
        code = CODE_REMAP.get(raw_code, raw_code)
        if any(code.startswith(p) for p in CODE_SKIP_PREFIXES):
            continue

        qty_val = row[qty_col] if qty_col < len(row) else None
        try:
            qty = int(qty_val) if qty_val is not None else 0
        except (ValueError, TypeError):
            qty = 0
        if qty <= 0:
            continue

        price_val = row[price_col] if price_col >= 0 and price_col < len(row) else None
        try:
            price = int(price_val) if price_val is not None else 0
        except (ValueError, TypeError):
            price = 0

        name = str(row[name_col]).strip() if name_col >= 0 and name_col < len(row) and row[name_col] is not None else ''
        store = str(row[store_col]).strip() if store_col >= 0 and store_col < len(row) and row[store_col] is not None else ''
        scat = str(row[scat_col]).strip() if scat_col >= 0 and scat_col < len(row) and row[scat_col] is not None else ''
        search_code = str(row[sc_col]).strip() if sc_col >= 0 and sc_col < len(row) and row[sc_col] is not None else ''

        category = get_code_cat(code) or (map_category(scat) if scat else 'その他')

        if ym not in monthly_data:
            monthly_data[ym] = {'by_store': {}, 'products': {}, 'date_range': None}
        processed_yms.add(ym)

        up = monthly_data[ym]
        if store not in up['by_store']:
            up['by_store'][store] = {}
        bs = up['by_store'][store]
        if category not in bs:
            bs[category] = {}
        entry = bs[category]
        if code not in entry:
            entry[code] = [0, 0]
        entry[code][0] += qty
        entry[code][1] += price * qty

        if code not in up['products'] or (name and not up['products'][code][0]):
            up['products'][code] = [name, price, search_code]
        elif search_code and not up['products'][code][2]:
            up['products'][code][2] = search_code

        dr = up['date_range']
        if dr is None:
            up['date_range'] = [date_str, date_str]
        else:
            if date_str < dr[0]:
                dr[0] = date_str
            if date_str > dr[1]:
                dr[1] = date_str

    return monthly_data, processed_yms


def merge_into_preloaded(preloaded, monthly_data):
    """
    monthly_data の年月を PRELOADED に上書きマージする。
    同じ年月のデータはクリアしてから新規セット。他の年月は保持。
    products は全マージ（新データ優先）。
    """
    by_store = preloaded.setdefault('by_store', {})

    # date_ranges も更新
    date_ranges = preloaded.setdefault('date_ranges', {})

    for ym, data in monthly_data.items():
        # その年月のデータを上書き
        by_store[ym] = data['by_store']
        # date_ranges: 年単位で管理
        year = ym[:4]
        dr = data.get('date_range')
        if dr:
            if year not in date_ranges or date_ranges[year] is None:
                date_ranges[year] = list(dr)
            else:
                if dr[0] < date_ranges[year][0]:
                    date_ranges[year][0] = dr[0]
                if dr[1] > date_ranges[year][1]:
                    date_ranges[year][1] = dr[1]

    # products は全マージ（新データ優先）
    products = preloaded.setdefault('products', {})
    for ym_data in monthly_data.values():
        for code, info in ym_data.get('products', {}).items():
            if code not in products or (info[0] and not products[code][0]):
                products[code] = info
            elif info[2] and not products[code][2]:
                products[code][2] = info[2]

    return preloaded


def write_preloaded(html_content, preloaded):
    """index.html の const PRELOADED= 行を新しい JSON で置き換える"""
    new_line = f'const PRELOADED={json.dumps(preloaded, ensure_ascii=False, separators=(",", ":"))};'
    new_content, count = re.subn(
        r'const PRELOADED=\{.*?\};?\s*\n',
        new_line + '\n',
        html_content,
        count=1,
        flags=re.DOTALL,
    )
    if count == 0:
        raise ValueError("index.html の const PRELOADED= 行が置換できませんでした")
    return new_content


def main():
    if len(sys.argv) < 2:
        print("使い方: python item-data/update_sales.py <Excelファイル1> [<Excelファイル2> ...]")
        sys.exit(1)

    excel_files = sys.argv[1:]
    for f in excel_files:
        if not os.path.exists(f):
            print(f"エラー: ファイルが見つかりません: {f}")
            sys.exit(1)

    print(f"index.html を読み込み中: {INDEX_HTML}")
    preloaded, html_content = load_preloaded(INDEX_HTML)

    all_processed_yms = set()
    for excel_file in excel_files:
        monthly_data, processed_yms = process_excel(excel_file)
        if not processed_yms:
            print(f"  警告: 有効なデータがありませんでした: {excel_file}")
            continue
        merge_into_preloaded(preloaded, monthly_data)
        all_processed_yms.update(processed_yms)

    if not all_processed_yms:
        print("エラー: 処理できるデータがありませんでした")
        sys.exit(1)

    print("index.html を更新中...")
    new_html = write_preloaded(html_content, preloaded)

    with open(INDEX_HTML, 'w', encoding='utf-8') as f:
        f.write(new_html)

    yms_sorted = sorted(all_processed_yms)
    print(f"完了: {len(yms_sorted)} 件の期間データを更新しました")
    print(f"  対象年月: {yms_sorted[0]} 〜 {yms_sorted[-1]}")


if __name__ == '__main__':
    main()
